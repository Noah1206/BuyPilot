"""
Competitor Analysis API
Simplified workflow: Keyword → Search Naver → Top 3 products → Calculate costs → Export
"""

from flask import Blueprint, request, jsonify, send_file
from datetime import datetime
import logging
import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from connectors.naver_shopping_scraper import get_naver_shopping_scraper
from utils.pricing import batch_calculate_costs
from utils.shipping import get_all_rates

bp = Blueprint('competitor', __name__)
logger = logging.getLogger(__name__)


@bp.route('/competitor/analyze', methods=['POST'])
def analyze_competitor():
    """
    Analyze competitor products by keyword

    Workflow:
    1. Search Naver Shopping by keyword
    2. Get top 3 products (sorted by sales)
    3. Calculate costs with 20% margin
    4. Return analysis results with product URLs

    Body: {
        keyword: string  # Search keyword (e.g., "캠핑용 화로", "공구")
    }

    Returns: {
        ok: boolean,
        data: {
            keyword: string,
            total_found: number,
            products: [
                {
                    title: string,
                    price: number,
                    url: string,  # Product URL link
                    category: string,
                    estimated_weight: number,
                    cost_price: number,
                    shipping_cost: number,
                    total_cost: number,
                    profit: number,
                    profit_rate: number,
                    recommendation: string,
                    image_url: string,
                    store_name: string,
                    rank: number
                }
            ]
        }
    }
    """
    try:
        data = request.get_json(force=True)

        keyword = data.get('keyword')
        if not keyword:
            return jsonify({
                'ok': False,
                'error': {
                    'code': 'VALIDATION_ERROR',
                    'message': 'Missing required field: keyword',
                    'details': {}
                }
            }), 400

        logger.info(f"🔍 Starting competitor analysis for keyword: {keyword}")

        # Step 1: Search Naver Shopping and get top 3 products
        scraper = get_naver_shopping_scraper()
        top_products = scraper.get_top_products(keyword, top_n=3)

        if not top_products:
            return jsonify({
                'ok': False,
                'error': {
                    'code': 'NO_PRODUCTS_FOUND',
                    'message': f'No products found for keyword: {keyword}',
                    'details': {}
                }
            }), 404

        # Step 2: Calculate costs for top 3 products
        analyzed_products = batch_calculate_costs(top_products)

        logger.info(f"✅ Analyzed {len(analyzed_products)} products for keyword: {keyword}")

        return jsonify({
            'ok': True,
            'data': {
                'keyword': keyword,
                'total_found': len(top_products),
                'products': analyzed_products,
                'analyzed_at': datetime.now().isoformat()
            }
        }), 200

    except Exception as e:
        logger.error(f"❌ Competitor analysis failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'ok': False,
            'error': {
                'code': 'ANALYSIS_ERROR',
                'message': 'Failed to analyze competitor',
                'details': {'error': str(e)}
            }
        }), 500


@bp.route('/competitor/export', methods=['POST'])
def export_analysis():
    """
    Export competitor analysis to Excel

    3 sheets:
    1. Product Analysis - detailed cost breakdowns with URLs
    2. Shipping Rates - Aceship GOLD rate table
    3. Summary - analysis overview

    Body: {
        products: array,  # Analyzed products from /analyze endpoint
        keyword: string
    }

    Returns: Excel file download
    """
    try:
        data = request.get_json(force=True)

        products = data.get('products', [])
        keyword = data.get('keyword', '')

        if not products:
            return jsonify({
                'ok': False,
                'error': {
                    'code': 'VALIDATION_ERROR',
                    'message': 'Missing required field: products',
                    'details': {}
                }
            }), 400

        logger.info(f"📊 Exporting analysis for {len(products)} products (keyword: {keyword})")

        # Create workbook
        wb = Workbook()

        # Sheet 1: Product Analysis
        ws1 = wb.active
        ws1.title = "Product Analysis"

        # Headers
        headers = [
            '순위', '상품명', '스토어', '카테고리', '경쟁사 가격',
            '예상 무게(kg)', '원가', '배송비', '총 비용',
            '이익', '마진율(%)', '추천', '상품 URL'
        ]

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(headers, start=1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Product data
        for idx, product in enumerate(products, start=2):
            ws1.cell(row=idx, column=1, value=product.get('rank', idx-1))
            ws1.cell(row=idx, column=2, value=product.get('title', ''))
            ws1.cell(row=idx, column=3, value=product.get('store_name', ''))
            ws1.cell(row=idx, column=4, value=product.get('category', ''))
            ws1.cell(row=idx, column=5, value=product.get('selling_price', 0))
            ws1.cell(row=idx, column=6, value=product.get('estimated_weight', 0))
            ws1.cell(row=idx, column=7, value=product.get('cost_price', 0))
            ws1.cell(row=idx, column=8, value=product.get('shipping_cost', 0))
            ws1.cell(row=idx, column=9, value=product.get('total_cost', 0))
            ws1.cell(row=idx, column=10, value=product.get('profit', 0))
            ws1.cell(row=idx, column=11, value=product.get('profit_rate', 0))
            ws1.cell(row=idx, column=12, value=product.get('recommendation', ''))
            # Add clickable URL link
            url_cell = ws1.cell(row=idx, column=13, value=product.get('url', ''))
            url_cell.hyperlink = product.get('url', '')
            url_cell.font = Font(color='0000FF', underline='single')

        # Adjust column widths
        ws1.column_dimensions['A'].width = 6
        ws1.column_dimensions['B'].width = 40
        ws1.column_dimensions['C'].width = 20
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 12
        ws1.column_dimensions['F'].width = 14
        ws1.column_dimensions['G'].width = 12
        ws1.column_dimensions['H'].width = 12
        ws1.column_dimensions['I'].width = 12
        ws1.column_dimensions['J'].width = 12
        ws1.column_dimensions['K'].width = 12
        ws1.column_dimensions['L'].width = 30
        ws1.column_dimensions['M'].width = 50

        # Sheet 2: Shipping Rates
        ws2 = wb.create_sheet(title="Shipping Rates")

        ws2.cell(row=1, column=1, value='Aceship GOLD 배송비표').font = Font(bold=True, size=14)

        rate_headers = ['무게(kg)', '배송비(원)']
        for col, header in enumerate(rate_headers, start=1):
            cell = ws2.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        rates = get_all_rates()
        for idx, (weight, cost) in enumerate(sorted(rates.items()), start=4):
            ws2.cell(row=idx, column=1, value=weight)
            ws2.cell(row=idx, column=2, value=cost)

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 15

        # Sheet 3: Summary
        ws3 = wb.create_sheet(title="Summary")

        ws3.cell(row=1, column=1, value='경쟁사 분석 요약').font = Font(bold=True, size=14)

        ws3.cell(row=3, column=1, value='검색 키워드:').font = Font(bold=True)
        ws3.cell(row=3, column=2, value=keyword)

        ws3.cell(row=4, column=1, value='분석 상품 수:').font = Font(bold=True)
        ws3.cell(row=4, column=2, value=len(products))

        ws3.cell(row=5, column=1, value='분석 일시:').font = Font(bold=True)
        ws3.cell(row=5, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        # Profitability summary
        ws3.cell(row=7, column=1, value='수익성 분석').font = Font(bold=True, size=12)

        avg_profit_rate = sum(p.get('profit_rate', 0) for p in products) / len(products) if products else 0
        avg_profit = sum(p.get('profit', 0) for p in products) / len(products) if products else 0

        ws3.cell(row=8, column=1, value='평균 마진율:').font = Font(bold=True)
        ws3.cell(row=8, column=2, value=f"{avg_profit_rate:.2f}%")

        ws3.cell(row=9, column=1, value='평균 이익:').font = Font(bold=True)
        ws3.cell(row=9, column=2, value=f"{int(avg_profit):,}원")

        profitable_count = sum(1 for p in products if p.get('is_profitable', False))
        ws3.cell(row=10, column=1, value='수익성 있는 상품:').font = Font(bold=True)
        ws3.cell(row=10, column=2, value=f"{profitable_count}/{len(products)}")

        # Top products with URLs
        ws3.cell(row=12, column=1, value='상위 제품 URL').font = Font(bold=True, size=12)

        for idx, product in enumerate(products, start=13):
            rank_cell = ws3.cell(row=idx, column=1, value=f"TOP {product.get('rank', idx-12)}:")
            rank_cell.font = Font(bold=True)

            url_cell = ws3.cell(row=idx, column=2, value=product.get('url', ''))
            url_cell.hyperlink = product.get('url', '')
            url_cell.font = Font(color='0000FF', underline='single')

        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 60

        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()

        logger.info(f"✅ Excel exported: {temp_file.name}")

        # Send file
        filename = f"competitor_analysis_{keyword}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = send_file(
            temp_file.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

        # Clean up temp file after sending
        @response.call_on_close
        def cleanup():
            try:
                os.unlink(temp_file.name)
            except:
                pass

        return response

    except Exception as e:
        logger.error(f"❌ Excel export failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'ok': False,
            'error': {
                'code': 'EXPORT_ERROR',
                'message': 'Failed to export Excel',
                'details': {'error': str(e)}
            }
        }), 500
