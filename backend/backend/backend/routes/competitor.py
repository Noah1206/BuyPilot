"""
Competitor Analysis API
Simplified workflow: Keyword â†’ Search Naver â†’ Top 3 products â†’ Calculate costs â†’ Export
"""

from flask import Blueprint, request, jsonify, send_file
from datetime import datetime
import logging
import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from connectors.naver_shopping_api import get_shopping_api
from utils.pricing import batch_calculate_costs
from utils.shipping import get_all_rates

bp = Blueprint('competitor', __name__)
logger = logging.getLogger(__name__)


@bp.route('/competitor/analyze', methods=['POST'])
def analyze_competitor():
    """
    Analyze competitor products by keyword

    Workflow:
    1. Search Naver Shopping by keyword
    2. Get top 3 products (sorted by sales)
    3. Calculate costs with 20% margin
    4. Return analysis results with product URLs

    Body: {
        keyword: string  # Search keyword (e.g., "ìº í•‘ìš© í™”ë¡œ", "ê³µêµ¬")
        count: number    # Number of products to analyze (default: 3, max: 100)
    }

    Returns: {
        ok: boolean,
        data: {
            keyword: string,
            total_found: number,
            products: [
                {
                    title: string,
                    price: number,
                    url: string,  # Product URL link
                    category: string,
                    estimated_weight: number,
                    cost_price: number,
                    shipping_cost: number,
                    total_cost: number,
                    profit: number,
                    profit_rate: number,
                    recommendation: string,
                    image_url: string,
                    store_name: string,
                    rank: number
                }
            ]
        }
    }
    """
    try:
        data = request.get_json(force=True)

        keyword = data.get('keyword')
        count = data.get('count', 3)  # Default to 3 if not provided

        if not keyword:
            return jsonify({
                'ok': False,
                'error': {
                    'code': 'VALIDATION_ERROR',
                    'message': 'Missing required field: keyword',
                    'details': {}
                }
            }), 400

        # Validate count (1-9999)
        try:
            count = int(count)
            if count < 1 or count > 9999:
                count = min(max(count, 1), 9999)  # Clamp to 1-9999
        except (ValueError, TypeError):
            count = 3  # Default if invalid

        logger.info(f"ðŸ” Starting competitor analysis for keyword: {keyword}, count: {count}")

        # Step 1: Search Naver Shopping API (max 100 per call, need multiple calls for > 100)
        api = get_shopping_api()
        top_products = []

        # Naver API allows max 100 results per call, so we need multiple calls
        if count <= 100:
            top_products = api.search_popular_products(keyword=keyword, max_products=count)
        else:
            # Make multiple API calls to get more than 100 products
            calls_needed = (count + 99) // 100  # Ceiling division
            for call_num in range(calls_needed):
                start_idx = call_num * 100 + 1
                fetch_count = min(100, count - call_num * 100)

                try:
                    batch = api.search_products(keyword, display=fetch_count, start=start_idx)
                    top_products.extend(batch)
                    logger.info(f"   Fetched batch {call_num + 1}/{calls_needed}: {len(batch)} products")
                except Exception as e:
                    logger.warning(f"   Failed to fetch batch {call_num + 1}: {str(e)}")
                    break

        top_products = top_products[:count]  # Ensure exact count

        if not top_products:
            return jsonify({
                'ok': False,
                'error': {
                    'code': 'NO_PRODUCTS_FOUND',
                    'message': f'No products found for keyword: {keyword}',
                    'details': {}
                }
            }), 404

        # Step 2: Transform API format to match our internal format
        standardized_products = []
        for product in top_products:
            standardized_products.append({
                'title': product['title'],
                'price': product['price'],
                'url': product['product_url'],
                'image_url': product['image_url'],
                'store_name': product['mall_name'],
                'rank': product['rank']
            })

        # Step 3: Calculate costs for top 3 products
        analyzed_products = batch_calculate_costs(standardized_products)

        logger.info(f"âœ… Analyzed {len(analyzed_products)} products for keyword: {keyword}")

        return jsonify({
            'ok': True,
            'data': {
                'keyword': keyword,
                'total_found': len(top_products),
                'products': analyzed_products,
                'analyzed_at': datetime.now().isoformat()
            }
        }), 200

    except Exception as e:
        logger.error(f"âŒ Competitor analysis failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'ok': False,
            'error': {
                'code': 'ANALYSIS_ERROR',
                'message': 'Failed to analyze competitor',
                'details': {'error': str(e)}
            }
        }), 500


@bp.route('/competitor/export', methods=['POST'])
def export_analysis():
    """
    Export competitor analysis to Excel

    3 sheets:
    1. Product Analysis - detailed cost breakdowns with URLs
    2. Shipping Rates - Aceship GOLD rate table
    3. Summary - analysis overview

    Body: {
        products: array,  # Analyzed products from /analyze endpoint
        keyword: string
    }

    Returns: Excel file download
    """
    try:
        data = request.get_json(force=True)

        products = data.get('products', [])
        keyword = data.get('keyword', '')

        if not products:
            return jsonify({
                'ok': False,
                'error': {
                    'code': 'VALIDATION_ERROR',
                    'message': 'Missing required field: products',
                    'details': {}
                }
            }), 400

        logger.info(f"ðŸ“Š Exporting analysis for {len(products)} products (keyword: {keyword})")

        # Create workbook
        wb = Workbook()

        # Sheet 1: Product Analysis
        ws1 = wb.active
        ws1.title = "Product Analysis"

        # Headers
        headers = [
            'ìˆœìœ„', 'ìƒí’ˆëª…', 'ìŠ¤í† ì–´', 'ì¹´í…Œê³ ë¦¬', 'ê²½ìŸì‚¬ ê°€ê²©',
            'ì˜ˆìƒ ë¬´ê²Œ(kg)', 'ì›ê°€', 'ë°°ì†¡ë¹„', 'ì´ ë¹„ìš©',
            'ì´ìµ', 'ë§ˆì§„ìœ¨(%)', 'ì¶”ì²œ', 'ìƒí’ˆ URL'
        ]

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(headers, start=1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Product data
        for idx, product in enumerate(products, start=2):
            ws1.cell(row=idx, column=1, value=product.get('rank', idx-1))
            ws1.cell(row=idx, column=2, value=product.get('title', ''))
            ws1.cell(row=idx, column=3, value=product.get('store_name', ''))
            ws1.cell(row=idx, column=4, value=product.get('category', ''))
            ws1.cell(row=idx, column=5, value=product.get('selling_price', 0))
            ws1.cell(row=idx, column=6, value=product.get('estimated_weight', 0))
            ws1.cell(row=idx, column=7, value=product.get('cost_price', 0))
            ws1.cell(row=idx, column=8, value=product.get('shipping_cost', 0))
            ws1.cell(row=idx, column=9, value=product.get('total_cost', 0))
            ws1.cell(row=idx, column=10, value=product.get('profit', 0))
            ws1.cell(row=idx, column=11, value=product.get('profit_rate', 0))
            ws1.cell(row=idx, column=12, value=product.get('recommendation', ''))
            # Add clickable URL link
            url_cell = ws1.cell(row=idx, column=13, value=product.get('url', ''))
            url_cell.hyperlink = product.get('url', '')
            url_cell.font = Font(color='0000FF', underline='single')

        # Adjust column widths
        ws1.column_dimensions['A'].width = 6
        ws1.column_dimensions['B'].width = 40
        ws1.column_dimensions['C'].width = 20
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 12
        ws1.column_dimensions['F'].width = 14
        ws1.column_dimensions['G'].width = 12
        ws1.column_dimensions['H'].width = 12
        ws1.column_dimensions['I'].width = 12
        ws1.column_dimensions['J'].width = 12
        ws1.column_dimensions['K'].width = 12
        ws1.column_dimensions['L'].width = 30
        ws1.column_dimensions['M'].width = 50

        # Sheet 2: Shipping Rates
        ws2 = wb.create_sheet(title="Shipping Rates")

        ws2.cell(row=1, column=1, value='Aceship GOLD ë°°ì†¡ë¹„í‘œ').font = Font(bold=True, size=14)

        rate_headers = ['ë¬´ê²Œ(kg)', 'ë°°ì†¡ë¹„(ì›)']
        for col, header in enumerate(rate_headers, start=1):
            cell = ws2.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        rates = get_all_rates()
        for idx, (weight, cost) in enumerate(sorted(rates.items()), start=4):
            ws2.cell(row=idx, column=1, value=weight)
            ws2.cell(row=idx, column=2, value=cost)

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 15

        # Sheet 3: Summary
        ws3 = wb.create_sheet(title="Summary")

        ws3.cell(row=1, column=1, value='ê²½ìŸì‚¬ ë¶„ì„ ìš”ì•½').font = Font(bold=True, size=14)

        ws3.cell(row=3, column=1, value='ê²€ìƒ‰ í‚¤ì›Œë“œ:').font = Font(bold=True)
        ws3.cell(row=3, column=2, value=keyword)

        ws3.cell(row=4, column=1, value='ë¶„ì„ ìƒí’ˆ ìˆ˜:').font = Font(bold=True)
        ws3.cell(row=4, column=2, value=len(products))

        ws3.cell(row=5, column=1, value='ë¶„ì„ ì¼ì‹œ:').font = Font(bold=True)
        ws3.cell(row=5, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        # Profitability summary
        ws3.cell(row=7, column=1, value='ìˆ˜ìµì„± ë¶„ì„').font = Font(bold=True, size=12)

        avg_profit_rate = sum(p.get('profit_rate', 0) for p in products) / len(products) if products else 0
        avg_profit = sum(p.get('profit', 0) for p in products) / len(products) if products else 0

        ws3.cell(row=8, column=1, value='í‰ê·  ë§ˆì§„ìœ¨:').font = Font(bold=True)
        ws3.cell(row=8, column=2, value=f"{avg_profit_rate:.2f}%")

        ws3.cell(row=9, column=1, value='í‰ê·  ì´ìµ:').font = Font(bold=True)
        ws3.cell(row=9, column=2, value=f"{int(avg_profit):,}ì›")

        profitable_count = sum(1 for p in products if p.get('is_profitable', False))
        ws3.cell(row=10, column=1, value='ìˆ˜ìµì„± ìžˆëŠ” ìƒí’ˆ:').font = Font(bold=True)
        ws3.cell(row=10, column=2, value=f"{profitable_count}/{len(products)}")

        # Top products with URLs
        ws3.cell(row=12, column=1, value='ìƒìœ„ ì œí’ˆ URL').font = Font(bold=True, size=12)

        for idx, product in enumerate(products, start=13):
            rank_cell = ws3.cell(row=idx, column=1, value=f"TOP {product.get('rank', idx-12)}:")
            rank_cell.font = Font(bold=True)

            url_cell = ws3.cell(row=idx, column=2, value=product.get('url', ''))
            url_cell.hyperlink = product.get('url', '')
            url_cell.font = Font(color='0000FF', underline='single')

        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 60

        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()

        logger.info(f"âœ… Excel exported: {temp_file.name}")

        # Send file
        filename = f"competitor_analysis_{keyword}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = send_file(
            temp_file.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

        # Clean up temp file after sending
        @response.call_on_close
        def cleanup():
            try:
                os.unlink(temp_file.name)
            except:
                pass

        return response

    except Exception as e:
        logger.error(f"âŒ Excel export failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'ok': False,
            'error': {
                'code': 'EXPORT_ERROR',
                'message': 'Failed to export Excel',
                'details': {'error': str(e)}
            }
        }), 500
